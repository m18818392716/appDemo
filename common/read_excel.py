from openpyxl import load_workbook

class readExcel():

    def __init__(self, file_name, sheet_name):
        self.file_name = file_name
        self.sheet_name = sheet_name

    def get_data1(self):
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]
        test_data = []  # 数据存储为列表格式

        for i in range(1, sheet.max_row):
            sub_data = {}  # 数据存储为字典格式
            sub_data['url'] = sheet.cell(i + 1, 2).value
            sub_data['data'] = sheet.cell(i + 1, 3).value
            sub_data['method'] = sheet.cell(i + 1, 4).value

            test_data.append(sub_data)

        return test_data  # 返回获取到的数据

    def get_data2(self):
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]
        cases = []  # 新建一个空列表，用于存放读取出的数据
        titles = []  # 新建一个空列表，用于存放读取到的表头，也就是1,2,3,4,5这一行
        rows = list(sheet.rows)  # 这里是将指定的sheet页中所有存在数据的行全都读取出来，转换成列表类型存放，方便我们进行遍历
        for row in rows[0]:
            titles.append(row.value)  # 将每一个表格的value值，也就是我们需要的数据，添加的空列表中。
        # 这里是遍历除了表头一行，剩下的所有行
        for row in rows[1:]:
            data = []
            for r in row:  # 遍历每一行的每一个表格
                data.append(r.value)
            data_zip = dict(zip(titles, data))  # 然后将每一行读取到的测试数据，和表头进行打包成一个字典的形式存放。
            cases.append(data_zip)  # 将所有测试数据添加到一个空列表中
        return cases


    def get_data(self, driver):

        wb = load_workbook('test_data.xlsx')  # 加载测试数据表格
        # sheets = wb.get_sheet_names() # 获取当前sheet名称

        # sheet = wb._add_sheet()  # 新建sheet对象

        # sheet.title  # 查看新建sheet对象的标题
        # sheet['A1'].value  # 查看sheet页中A1单元格的数值


        sheet = wb.get_sheet_by_name('Sheet1')

        # 通过枚举函数来循环处理所有的数据行，注意去除标题行
        for i, _ in enumerate(list(sheet.rows)[:-1]):
            id = sheet['C' + str(i + 2)].value  # 读取测试数据字段
            driver.find_element_by_id(id).click()  # 点击指定的id的按键

            # 元素定位到公式显示区域，获取此元素的文本值作为实际结果
            text = driver.find_element_by_id('com.android.calculator2:id/formula').text
            sheet['E' + str(i + 2)].value = text  # 给E列赋值，实际返回结果
            expected_result = sheet['D' + str(i + 2)].value  # 获取预期结果字段

        # 判断实际结果和预期结果是否一致,一致的话标记pass，否则标记fail。
        if (text == expected_result):
            sheet['F' + str(i + 2)].value = 'pass'
        else:
            sheet['F' + str(i + 2)].value = 'fail'

        wb.save('updatedTestData.xlsx')  # 另存为表格文档
        driver.quit()

    def open_excel(self):
        self.wb = load_workbook(filename=self.file_name)
        self.sh = self.wb['data']
    def write_excel(self, row, column, value):
        '''

        :param row: 指定写入的行
        :param column: 指定写入的列
        :param value: 指定写入的数据
        :return:
        '''
        # 打开
        self.open_excel()
        # 写入
        self.sh.cell(row=row, column=column, value=value)
        # 保存
        self.wb.save(self.file_name)

if __name__ == "__main__":
    filepath = "../data/test_data.xlsx"
    sheetName = "Sheet1"
    data = readExcel(filepath, sheetName)
    print(data.get_data2())